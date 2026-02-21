"""
AQI Agreement & Performance Charts - Excel Native Charts
=========================================================
Generates a single .xlsx with 8 sheets, each containing:
  - Raw data columns (Sample, Actual AQI, Predicted AQI)
  - Agreement Plot (Scatter chart) - editable Excel chart
  - Performance Plot (Line chart) - editable Excel chart
  - Metrics summary (CC, R2, RMSE, MAE)

All charts are NATIVE Excel charts - fully editable in Excel.
All data comes directly from the source Excel file.
"""

import openpyxl
from openpyxl.chart import ScatterChart, LineChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.trendline import Trendline
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.legend import Legend
from openpyxl.drawing.line import LineProperties, LineEndProperties
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
import numpy as np
import os

# ============================================================
#  PATHS
# ============================================================
SRC_EXCEL = r"C:\Users\HP\OneDrive\Desktop\research papers 2026\AQI PAPER\AGREEMENT GRAPHS\M5P TRAINING 1.xlsx"
OUT_DIR   = r"C:\Users\HP\OneDrive\Desktop\research papers 2026\AQI PAPER\AGREEMENT GRAPHS\output_plots"
OUT_EXCEL = os.path.join(OUT_DIR, "AQI_Charts_Editable.xlsx")
os.makedirs(OUT_DIR, exist_ok=True)

# ============================================================
#  SHEET METADATA
# ============================================================
SHEET_INFO = [
    ('Sheet1', 'M5P',           'Training', True),   # has inst# col
    ('Sheet2', 'M5P',           'Testing',  False),
    ('Sheet3', 'Random Forest', 'Training', False),
    ('Sheet4', 'Random Forest', 'Testing',  False),
    ('Sheet5', 'Random Tree',   'Training', False),
    ('Sheet6', 'REP Tree',      'Testing',  False),
    ('Sheet7', 'REP Tree',      'Training', False),
    ('Sheet8', 'REP Tree',      'Testing',  False),
]

# ============================================================
#  STYLES
# ============================================================
HEADER_FONT   = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL   = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
METRIC_FONT   = Font(name='Calibri', bold=True, size=11, color='1F4E79')
METRIC_VAL    = Font(name='Calibri', size=11)
TITLE_FONT    = Font(name='Calibri', bold=True, size=14, color='1F4E79')
DATA_FONT     = Font(name='Calibri', size=10)
THIN_BORDER   = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Chart color palettes (hex without #)
SCATTER_COLORS = ['2166AC', 'D6604D', '4DAF4A', '984EA3', 'FF7F00', 'A65628', '377EB8', 'E41A1C']
LINE_ACTUAL    = ['1B7837', '2166AC', '1B9E77', '7570B3', '66A61E', 'A6761D', 'E41A1C', '4DAF4A']
LINE_PRED      = ['762A83', 'B2182B', 'D95F02', 'E7298A', 'E6AB02', '666666', '377EB8', '984EA3']

# Marker symbols for scatter (openpyxl marker codes)
MARKER_STYLES = ['circle', 'square', 'triangle', 'diamond', 'circle', 'square', 'triangle', 'diamond']


def compute_metrics(actual, predicted):
    """Compute CC, R2, RMSE, MAE, NSE."""
    actual = np.array(actual, dtype=float)
    predicted = np.array(predicted, dtype=float)
    n = len(actual)
    cc = float(np.corrcoef(actual, predicted)[0, 1])
    ss_res = float(np.sum((actual - predicted) ** 2))
    ss_tot = float(np.sum((actual - np.mean(actual)) ** 2))
    r2 = 1 - ss_res / ss_tot if ss_tot != 0 else 0
    mae = float(np.mean(np.abs(actual - predicted)))
    rmse = float(np.sqrt(np.mean((actual - predicted) ** 2)))
    nse = r2  # same formula
    return {'CC': cc, 'R2': r2, 'MAE': mae, 'RMSE': rmse, 'NSE': nse, 'N': n}


def create_agreement_chart(ws, data_rows, col_actual, col_pred, color_hex, marker, sheet_label, anchor_cell):
    """
    Create a Scatter chart (Agreement Plot) with:
    - Data points, 1:1 reference line, trendline, metrics
    """
    chart = ScatterChart()
    chart.title = f"Agreement Plot: {sheet_label}"
    chart.x_axis.title = "Observed AQI"
    chart.y_axis.title = "Predicted AQI"
    chart.style = 2
    chart.width = 18
    chart.height = 14

    # Data series (actual vs predicted scatter)
    xvals = Reference(ws, min_col=col_actual, min_row=2, max_row=data_rows + 1)
    yvals = Reference(ws, min_col=col_pred, min_row=2, max_row=data_rows + 1)

    series = Series(yvals, xvals, title="Data Points")
    series.graphicalProperties = GraphicalProperties()
    series.graphicalProperties.line.noFill = True  # no connecting line
    series.marker = openpyxl.chart.marker.Marker(symbol=marker, size=4)
    series.marker.graphicalProperties = GraphicalProperties()
    series.marker.graphicalProperties.solidFill = color_hex
    series.marker.graphicalProperties.line.solidFill = color_hex

    # Add trendline (linear regression)
    trend = Trendline(trendlineType='linear', dispRSqr=True, dispEq=True)
    series.trendline = trend

    chart.series.append(series)

    # 1:1 line - use the min/max reference line data from columns
    # We wrote 1:1 line data in specific columns
    col_11x = col_pred + 2  # column for 1:1 X
    col_11y = col_11x + 1   # column for 1:1 Y
    xvals_11 = Reference(ws, min_col=col_11x, min_row=2, max_row=3)
    yvals_11 = Reference(ws, min_col=col_11y, min_row=2, max_row=3)

    series_11 = Series(yvals_11, xvals_11, title="1:1 Line")
    series_11.graphicalProperties = GraphicalProperties()
    series_11.graphicalProperties.line.solidFill = '000000'
    series_11.graphicalProperties.line.width = 15000  # EMU
    series_11.graphicalProperties.line.dashStyle = 'solid'
    series_11.marker = openpyxl.chart.marker.Marker(symbol='none')

    chart.series.append(series_11)

    # Axis formatting
    chart.x_axis.tickLblPos = 'low'
    chart.y_axis.tickLblPos = 'low'
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.majorGridlines = None
    chart.y_axis.majorGridlines = None

    # Legend
    chart.legend.position = 'b'

    ws.add_chart(chart, anchor_cell)
    return chart


def create_performance_chart(ws, data_rows, col_idx, col_actual, col_pred,
                              color_act, color_pred, sheet_label, anchor_cell):
    """
    Create a Line chart (Performance Plot) with:
    - Actual AQI (solid line with markers)
    - Predicted AQI (dashed line with markers)
    """
    chart = LineChart()
    chart.title = f"Performance Plot: {sheet_label}"
    chart.x_axis.title = "Sample Index"
    chart.y_axis.title = "AQI Value"
    chart.style = 2
    chart.width = 18
    chart.height = 14

    # Category axis (sample index)
    cats = Reference(ws, min_col=col_idx, min_row=2, max_row=data_rows + 1)

    # Actual AQI series
    actual_data = Reference(ws, min_col=col_actual, min_row=1, max_row=data_rows + 1)
    chart.add_data(actual_data, titles_from_data=True)
    chart.set_categories(cats)

    # Predicted AQI series
    pred_data = Reference(ws, min_col=col_pred, min_row=1, max_row=data_rows + 1)
    chart.add_data(pred_data, titles_from_data=True)

    # Style actual series
    s0 = chart.series[0]
    s0.graphicalProperties.line.solidFill = color_act
    s0.graphicalProperties.line.width = 18000
    s0.marker = openpyxl.chart.marker.Marker(symbol='circle', size=3)
    s0.marker.graphicalProperties = GraphicalProperties()
    s0.marker.graphicalProperties.solidFill = color_act
    s0.marker.graphicalProperties.line.solidFill = color_act
    s0.smooth = False

    # Style predicted series
    s1 = chart.series[1]
    s1.graphicalProperties.line.solidFill = color_pred
    s1.graphicalProperties.line.width = 18000
    s1.graphicalProperties.line.dashStyle = 'dash'
    s1.marker = openpyxl.chart.marker.Marker(symbol='square', size=3)
    s1.marker.graphicalProperties = GraphicalProperties()
    s1.marker.graphicalProperties.solidFill = color_pred
    s1.marker.graphicalProperties.line.solidFill = color_pred
    s1.smooth = False

    # Axis formatting
    chart.x_axis.tickLblPos = 'low'
    chart.y_axis.tickLblPos = 'low'

    # Legend
    chart.legend.position = 'b'

    # Reduce clutter for large datasets - show fewer tick labels
    if data_rows > 200:
        chart.x_axis.tickLblSkip = max(1, data_rows // 10)
        chart.x_axis.tickMarkSkip = max(1, data_rows // 10)

    ws.add_chart(chart, anchor_cell)
    return chart


def main():
    print("=" * 60)
    print("  AQI Excel Charts Generator (Native Editable Charts)")
    print("  All data from source Excel - zero fabrication")
    print("=" * 60)

    # Load source
    src_wb = openpyxl.load_workbook(SRC_EXCEL, data_only=True)
    print(f"  Source: {SRC_EXCEL}")
    print(f"  Sheets: {src_wb.sheetnames}")

    # Create output workbook
    out_wb = openpyxl.Workbook()
    # Remove default sheet
    out_wb.remove(out_wb.active)

    all_metrics = []

    for i, (src_sheet, model, dtype, has_idx) in enumerate(SHEET_INFO):
        src_ws = src_wb[src_sheet]
        sheet_label = f"{model} ({dtype})"
        tab_name = f"{model[:8]}_{dtype[:5]}"  # Keep sheet name <= 31 chars
        print(f"\n  [{i+1}/8] {sheet_label} ({src_sheet})")

        # Read data from source
        actual_vals = []
        pred_vals = []

        # Determine column positions
        # Sheet1 has: inst#, actual, predicted, None, None
        # All others have: actual, predicted, None, label
        if has_idx:
            act_col_src, pred_col_src = 2, 3
        else:
            act_col_src, pred_col_src = 1, 2

        for r in range(2, src_ws.max_row + 1):
            a = src_ws.cell(r, act_col_src).value
            p = src_ws.cell(r, pred_col_src).value
            if a is not None and p is not None:
                try:
                    actual_vals.append(float(a))
                    pred_vals.append(float(p))
                except (ValueError, TypeError):
                    continue

        n = len(actual_vals)
        print(f"    Data points: {n}")

        if n == 0:
            print(f"    [SKIP] No valid data")
            continue

        # Compute metrics
        metrics = compute_metrics(actual_vals, pred_vals)
        metrics['Model'] = model
        metrics['Type'] = dtype
        metrics['Sheet'] = src_sheet
        all_metrics.append(metrics)
        print(f"    CC={metrics['CC']:.4f}  R2={metrics['R2']:.4f}  RMSE={metrics['RMSE']:.2f}  MAE={metrics['MAE']:.2f}")

        # Create worksheet in output
        ws = out_wb.create_sheet(title=tab_name)

        # ─── WRITE DATA COLUMNS ───
        # Col A: Sample Index
        # Col B: Actual AQI
        # Col C: Predicted AQI
        # Col D: (empty)
        # Col E: 1:1 Line X  (for agreement chart reference line)
        # Col F: 1:1 Line Y

        headers = ['Sample #', 'Actual AQI', 'Predicted AQI', '', '1:1 Line X', '1:1 Line Y']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER

        # Write data
        for r in range(n):
            ws.cell(row=r+2, column=1, value=r+1).font = DATA_FONT
            ws.cell(row=r+2, column=1).border = THIN_BORDER
            ws.cell(row=r+2, column=1).alignment = Alignment(horizontal='center')

            cell_a = ws.cell(row=r+2, column=2, value=round(actual_vals[r], 3))
            cell_a.font = DATA_FONT
            cell_a.border = THIN_BORDER
            cell_a.number_format = '0.000'

            cell_p = ws.cell(row=r+2, column=3, value=round(pred_vals[r], 3))
            cell_p.font = DATA_FONT
            cell_p.border = THIN_BORDER
            cell_p.number_format = '0.000'

        # 1:1 line reference data (min and max)
        all_v = actual_vals + pred_vals
        v_min = max(0, min(all_v) - 20)
        v_max = max(all_v) + 20
        ws.cell(row=2, column=5, value=round(v_min, 1))
        ws.cell(row=3, column=5, value=round(v_max, 1))
        ws.cell(row=2, column=6, value=round(v_min, 1))
        ws.cell(row=3, column=6, value=round(v_max, 1))

        # ─── WRITE METRICS (Col H-I) ───
        metrics_start_col = 8
        metric_labels = ['Metric', 'CC', 'R-squared', 'RMSE', 'MAE', 'NSE', 'N', '', 'Model', 'Dataset', 'Source Sheet']
        metric_values = ['Value',
                         round(metrics['CC'], 6),
                         round(metrics['R2'], 6),
                         round(metrics['RMSE'], 4),
                         round(metrics['MAE'], 4),
                         round(metrics['NSE'], 6),
                         metrics['N'],
                         '',
                         model,
                         dtype,
                         src_sheet]

        for r, (lbl, val) in enumerate(zip(metric_labels, metric_values), 1):
            cell_l = ws.cell(row=r, column=metrics_start_col, value=lbl)
            cell_v = ws.cell(row=r, column=metrics_start_col + 1, value=val)
            if r == 1:
                cell_l.font = HEADER_FONT
                cell_l.fill = HEADER_FILL
                cell_v.font = HEADER_FONT
                cell_v.fill = HEADER_FILL
            else:
                cell_l.font = METRIC_FONT
                cell_v.font = METRIC_VAL
                if isinstance(val, float):
                    cell_v.number_format = '0.000000'
            cell_l.border = THIN_BORDER
            cell_v.border = THIN_BORDER
            cell_l.alignment = Alignment(horizontal='left')
            cell_v.alignment = Alignment(horizontal='center')

        # Column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 3
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 3
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 16

        # ─── CREATE AGREEMENT CHART (Scatter) ───
        # Place chart at row 14 (below data area would be too far down for large datasets)
        # Instead, place charts to the right of the metrics
        agree_anchor = 'K2'
        create_agreement_chart(
            ws, n, col_actual=2, col_pred=3,
            color_hex=SCATTER_COLORS[i], marker=MARKER_STYLES[i],
            sheet_label=sheet_label, anchor_cell=agree_anchor
        )

        # ─── CREATE PERFORMANCE CHART (Line) ───
        perf_anchor = 'K30'
        create_performance_chart(
            ws, n, col_idx=1, col_actual=2, col_pred=3,
            color_act=LINE_ACTUAL[i], color_pred=LINE_PRED[i],
            sheet_label=sheet_label, anchor_cell=perf_anchor
        )

        print(f"    [OK] Sheet '{tab_name}' created with 2 editable charts")

    # ─── METRICS SUMMARY SHEET ───
    ws_sum = out_wb.create_sheet(title="Metrics Summary")
    sum_headers = ['Sheet', 'Model', 'Type', 'CC', 'R-squared', 'RMSE', 'MAE', 'NSE', 'N']
    for c, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    for r, m in enumerate(all_metrics, 2):
        vals = [m['Sheet'], m['Model'], m['Type'],
                round(m['CC'], 6), round(m['R2'], 6),
                round(m['RMSE'], 4), round(m['MAE'], 4),
                round(m['NSE'], 6), m['N']]
        for c, v in enumerate(vals, 1):
            cell = ws_sum.cell(row=r, column=c, value=v)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
            if isinstance(v, float):
                cell.number_format = '0.000000'

    # Adjust column widths
    for c, w in enumerate([12, 16, 10, 12, 12, 12, 12, 12, 8], 1):
        ws_sum.column_dimensions[get_column_letter(c)].width = w

    # Add a comparison bar chart for R2 values
    from openpyxl.chart import BarChart
    bar_chart = BarChart()
    bar_chart.type = 'col'
    bar_chart.title = 'R-squared Comparison Across Models'
    bar_chart.x_axis.title = 'Model'
    bar_chart.y_axis.title = 'R-squared'
    bar_chart.width = 20
    bar_chart.height = 12
    bar_chart.style = 2

    cats = Reference(ws_sum, min_col=2, min_row=2, max_row=len(all_metrics)+1)
    r2_data = Reference(ws_sum, min_col=5, min_row=1, max_row=len(all_metrics)+1)
    bar_chart.add_data(r2_data, titles_from_data=True)
    bar_chart.set_categories(cats)
    bar_chart.shape = 4

    # Color each bar
    if bar_chart.series:
        s = bar_chart.series[0]
        s.graphicalProperties.line.solidFill = '1F4E79'
        s.graphicalProperties.solidFill = '2166AC'

    ws_sum.add_chart(bar_chart, 'A12')

    # Add RMSE comparison chart
    rmse_chart = BarChart()
    rmse_chart.type = 'col'
    rmse_chart.title = 'RMSE Comparison Across Models'
    rmse_chart.x_axis.title = 'Model'
    rmse_chart.y_axis.title = 'RMSE'
    rmse_chart.width = 20
    rmse_chart.height = 12
    rmse_chart.style = 2

    rmse_data = Reference(ws_sum, min_col=6, min_row=1, max_row=len(all_metrics)+1)
    rmse_chart.add_data(rmse_data, titles_from_data=True)
    rmse_chart.set_categories(cats)

    if rmse_chart.series:
        s = rmse_chart.series[0]
        s.graphicalProperties.solidFill = 'D6604D'
        s.graphicalProperties.line.solidFill = '8B0000'

    ws_sum.add_chart(rmse_chart, 'L12')

    # CC comparison chart
    cc_chart = BarChart()
    cc_chart.type = 'col'
    cc_chart.title = 'Correlation Coefficient (CC) Comparison'
    cc_chart.x_axis.title = 'Model'
    cc_chart.y_axis.title = 'CC'
    cc_chart.width = 20
    cc_chart.height = 12
    cc_chart.style = 2

    cc_data = Reference(ws_sum, min_col=4, min_row=1, max_row=len(all_metrics)+1)
    cc_chart.add_data(cc_data, titles_from_data=True)
    cc_chart.set_categories(cats)

    if cc_chart.series:
        s = cc_chart.series[0]
        s.graphicalProperties.solidFill = '4DAF4A'
        s.graphicalProperties.line.solidFill = '1B7837'

    ws_sum.add_chart(cc_chart, 'A30')

    # ─── SAVE ───
    out_wb.save(OUT_EXCEL)
    fsize = os.path.getsize(OUT_EXCEL)
    print(f"\n{'=' * 60}")
    print(f"  Excel saved: {OUT_EXCEL}")
    print(f"  Size: {fsize:,} bytes ({fsize/1024:.0f} KB)")
    print(f"  Sheets: {len(out_wb.sheetnames)} ({', '.join(out_wb.sheetnames)})")
    print(f"  All charts are NATIVE Excel charts - fully editable!")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()

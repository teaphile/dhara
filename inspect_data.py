import openpyxl
import os

EXCEL = r'C:\Users\HP\OneDrive\Desktop\research papers 2026\AQI PAPER\AGREEMENT GRAPHS\M5P TRAINING 1.xlsx'
LOG = r'C:\Users\HP\OneDrive\Desktop\NICMAR PROJECT\data_inspect.log'

wb = openpyxl.load_workbook(EXCEL, data_only=True)
lines = []
lines.append(f"File: {EXCEL}")
lines.append(f"Sheets: {len(wb.sheetnames)}")
lines.append("")

for i, name in enumerate(wb.sheetnames):
    ws = wb[name]
    lines.append(f"=== Sheet {i+1}: [{name}] ===")
    lines.append(f"  Dimensions: {ws.max_row} rows x {ws.max_column} cols")
    # Header row
    hdr = [str(ws.cell(1, c).value) for c in range(1, ws.max_column+1)]
    lines.append(f"  Headers: {hdr}")
    # Sample rows 2-6
    for r in range(2, min(7, ws.max_row+1)):
        vals = []
        for c in range(1, ws.max_column+1):
            v = ws.cell(r, c).value
            vals.append(v)
        lines.append(f"  Row {r}: {vals}")
    # Last 2 rows
    if ws.max_row > 7:
        lines.append("  ...")
        for r in range(max(8, ws.max_row-1), ws.max_row+1):
            vals = []
            for c in range(1, ws.max_column+1):
                v = ws.cell(r, c).value
                vals.append(v)
            lines.append(f"  Row {r}: {vals}")
    # Count non-null in each col
    col_counts = []
    for c in range(1, ws.max_column+1):
        cnt = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, c).value is not None)
        col_counts.append(f"{hdr[c-1]}={cnt}")
    lines.append(f"  Non-null: {col_counts}")
    lines.append("")

with open(LOG, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f"Log written to {LOG}")
print(f"Size: {os.path.getsize(LOG)} bytes")
